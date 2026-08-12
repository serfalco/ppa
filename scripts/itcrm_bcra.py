"""
PPA — itcrm_bcra.py
Baja y parsea el Índice de Tipo de Cambio Real Multilateral del BCRA.

Por qué una planilla y no una API: el ITCRM no está en ninguna de las dos
APIs que usa el proyecto. La API del BCRA publica tipo de cambio minorista,
mayorista y de valuación contable, ninguno real ni multilateral. Y las series
de tipo de cambio real de datos.gob.ar son bilaterales (Canadá, China,
México, Uruguay, Vietnam, Chile) y cortaron todas el 28/01/2026. La planilla
oficial es la única publicación viva del índice.

Cómo se lee: el BCRA publica ITCRMSerie.xlsx con el índice diario base
17/12/2015=100 desde 1997, junto a los bilaterales por país en columnas
vecinas. El formato de la planilla cambió antes y va a volver a cambiar, así
que acá no hay coordenadas fijas: se busca la columna cuyo encabezado dice
ITCRM y la columna de fechas, y se leen las filas que tengan las dos cosas.
Si la planilla cambia de forma, el modo diagnóstico muestra qué vino.

Cómo se corre a mano (donde haya internet):
    python scripts/itcrm_bcra.py                # baja, parsea y resume
    python scripts/itcrm_bcra.py --diagnostico  # muestra la forma del archivo
"""

import io
import re
import sys
import unicodedata
from datetime import datetime, date

import requests

URL_ITCRM = "https://www.bcra.gob.ar/Pdfs/PublicacionesEstadisticas/ITCRMSerie.xlsx"

# Navegador real: el sitio del BCRA rechaza clientes que no lo parezcan.
HDRS = {
    "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                   "AppleWebKit/537.36 (KHTML, like Gecko) "
                   "Chrome/126.0 Safari/537.36"),
    "Accept": ("application/vnd.openxmlformats-officedocument."
               "spreadsheetml.sheet,application/vnd.ms-excel,*/*"),
}

# La planilla completa son ~7000 filas y varios megas: 60s y no 30.
TIMEOUT = 60

# Cuántas filas del principio se miran buscando el encabezado. El BCRA mete
# título, subtítulo y notas antes de la tabla, pero nunca tanto.
FILAS_ENCABEZADO = 25


def _sin_acentos(txt):
    """Normaliza para comparar encabezados sin depender de tildes ni mayúsculas."""
    txt = unicodedata.normalize("NFKD", str(txt or ""))
    return "".join(c for c in txt if not unicodedata.combining(c)).strip().lower()


def bajar(url=URL_ITCRM):
    """Devuelve el contenido de la planilla, o lanza con el motivo."""
    try:
        r = requests.get(url, headers=HDRS, timeout=TIMEOUT)
    except requests.exceptions.SSLError:
        # El BCRA tiene la cadena SSL incompleta cada tanto; el mismo caso
        # que ya contempla verificar_series.py contra su API.
        import urllib3
        urllib3.disable_warnings()
        r = requests.get(url, headers=HDRS, timeout=TIMEOUT, verify=False)
    r.raise_for_status()
    return r.content


def _abrir(contenido):
    import openpyxl
    return openpyxl.load_workbook(io.BytesIO(contenido), data_only=True,
                                  read_only=True)


def _es_fecha(v):
    return isinstance(v, (datetime, date))


def _a_iso(v):
    return (v.date() if isinstance(v, datetime) else v).isoformat()


def _numero(v):
    """El valor como float, o None si la celda no es un número usable.

    Las planillas del BCRA traen guiones y celdas vacías en los feriados, y
    a veces el número viene como texto con coma decimal.
    """
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    if isinstance(v, str):
        limpio = v.strip().replace(".", "").replace(",", ".")
        if re.fullmatch(r"-?\d+(\.\d+)?", limpio or ""):
            return float(limpio)
    return None


def ubicar_columnas(hoja):
    """Encuentra (fila_encabezado, col_fecha, col_itcrm) mirando el contenido.

    Se busca el encabezado que diga ITCRM; la columna de fechas es la primera
    de esa misma tabla que traiga fechas de verdad. Buscar por contenido y no
    por coordenadas es lo que hace que un cambio de layout no rompa nada.
    """
    filas = []
    for i, fila in enumerate(hoja.iter_rows(min_row=1, max_row=FILAS_ENCABEZADO,
                                            values_only=True), start=1):
        filas.append((i, fila))

    fila_enc = col_itcrm = None
    for i, fila in filas:
        for j, celda in enumerate(fila or ()):
            texto = _sin_acentos(celda)
            # "ITCRM" a secas, o el nombre desplegado. Se descartan los
            # bilaterales, que dicen ITCRB o nombran un país.
            if texto.startswith("itcrm") or (
                    "tipo de cambio real multilateral" in texto):
                fila_enc, col_itcrm = i, j
                break
        if fila_enc is not None:
            break

    if fila_enc is None:
        return None, None, None

    # La columna de fechas: la primera que tenga una fecha debajo del
    # encabezado. Casi siempre es la A, pero no se asume.
    col_fecha = None
    for fila in hoja.iter_rows(min_row=fila_enc + 1, max_row=fila_enc + 40,
                               values_only=True):
        for j, celda in enumerate(fila or ()):
            if _es_fecha(celda):
                col_fecha = j
                break
        if col_fecha is not None:
            break

    return fila_enc, col_fecha, col_itcrm


def parsear(contenido):
    """Devuelve (serie, meta). serie = [[fecha_iso, valor], ...] ascendente."""
    libro = _abrir(contenido)
    for nombre in libro.sheetnames:
        hoja = libro[nombre]
        fila_enc, col_fecha, col_itcrm = ubicar_columnas(hoja)
        if fila_enc is None or col_fecha is None:
            continue

        serie = []
        for fila in hoja.iter_rows(min_row=fila_enc + 1, values_only=True):
            if not fila or len(fila) <= max(col_fecha, col_itcrm):
                continue
            f, v = fila[col_fecha], _numero(fila[col_itcrm])
            if _es_fecha(f) and v is not None:
                serie.append([_a_iso(f), round(v, 4)])

        if serie:
            serie.sort(key=lambda p: p[0])
            meta = {"hoja": nombre, "fila_encabezado": fila_enc,
                    "col_fecha": col_fecha, "col_itcrm": col_itcrm,
                    "puntos": len(serie),
                    "desde": serie[0][0], "hasta": serie[-1][0]}
            return serie, meta

    raise ValueError("no encontré la columna del ITCRM en ninguna hoja")


def obtener(url=URL_ITCRM):
    """Baja y parsea de una. Devuelve (serie, meta)."""
    return parsear(bajar(url))


def _diagnostico():
    """Muestra la forma del archivo. Para cuando el parseo deja de andar."""
    print(f"Bajando {URL_ITCRM} …")
    contenido = bajar()
    print(f"   {len(contenido)} bytes")
    libro = _abrir(contenido)
    print(f"   hojas: {libro.sheetnames}")
    for nombre in libro.sheetnames:
        hoja = libro[nombre]
        print(f"\n=== hoja {nombre!r}")
        fila_enc, col_fecha, col_itcrm = ubicar_columnas(hoja)
        print(f"    encabezado detectado: fila={fila_enc} "
              f"col_fecha={col_fecha} col_itcrm={col_itcrm}")
        print("    primeras filas:")
        for i, fila in enumerate(hoja.iter_rows(min_row=1, max_row=12,
                                                values_only=True), start=1):
            celdas = [str(c)[:22] for c in (fila or ())[:8]]
            print(f"      {i:3d} | " + " | ".join(celdas))


def main():
    if "--diagnostico" in sys.argv[1:]:
        _diagnostico()
        return
    serie, meta = obtener()
    print(f"ITCRM · {meta['puntos']} puntos · "
          f"{meta['desde']} → {meta['hasta']}")
    print(f"   hoja {meta['hoja']!r}, encabezado en fila "
          f"{meta['fila_encabezado']}")
    print("   últimos valores:")
    for f, v in serie[-5:]:
        print(f"      {f}  {v}")


if __name__ == "__main__":
    main()
